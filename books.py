#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   douban.py
@Time    :   2021/02/18
@Author  :   HDUZN
@Version :   2.0
@Contact :   hduzn@vip.qq.com
@License :   (C)Copyright 2021-2022
@Desc    :   将Douban 读过的书的记录保存到douban.xlsx 文件和数据库中
             (注：豆瓣标记后，没写标签、没评分/星 会报错，没写评论没关系。)
             更新于2022/05/12，采用selenium4版本，跳过selenium的webdriver检测
'''

# here put the import lib
import douban_config, z_db
import time, os, re
import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import openpyxl

# 获取最大页数和每一页的模板链接
def get_max_page_num(wd):
    num_element = wd.find_element(By.XPATH, '//div[@class="paginator"]/a[last()]')
    max_page_num = int(num_element.text) # 21
    site = num_element.get_attribute('href')
    # num_site_list = [max_page_num, site]
    return [max_page_num, site]

# 获取所有页数的链接List
def get_page_site_list(max_page_num, model_site):
    # model_site: https://book.douban.com/people/[id]/collect?start=300&sort=time&rating=all&filter=all&mode=grid
    cop = '[0-9]'
    site_part2 = ''.join(re.findall(cop, model_site)) # '300'
    site_part1 = model_site.split(site_part2)[0] # https://book.douban.com/people/[id]/collect?start=
    site_part3 = model_site.split(site_part2)[1] # &sort=time&rating=all&filter=all&mode=grid

    site_list = []
    for i in range(max_page_num):
        site_part2_num = str(i * 15)
        page_site = site_part1 + site_part2_num + site_part3
        site_list.append(page_site)
    return site_list

# 获取读过的书的页面 wd
def get_readed_wd(site):
    option = webdriver.ChromeOptions()
    # option.add_experimental_option('detach', True) # 不自动关闭浏览器
    # option.add_experimental_option('excludeSwitches', ['enable-automation'])
    # option.add_experimental_option('useAutomationExtension', False)
    
    # 使用本地debugger模式，可以提前登录
    option.add_experimental_option('debuggerAddress','127.0.0.1:9222')
    wd = webdriver.Chrome(options=option)
    wd.implicitly_wait(10)
    wd.maximize_window()

    # print(wd.execute_script("return window.navigator.userAgent"))
    # wd.execute_cdp_cmd('Network.setUserAgentOverride', {"userAgent": user_agent})

    # # 默认window.navigator.webdriver=true
    # script = '''
    # Object.defineProperty(navigator, 'webdriver', {
    #     get: () => false
    # })
    # '''
    # wd.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {"source": script})

    wd.get(site)
    time.sleep(3)

    # 登录
    try:
        wd.find_element(By.CLASS_NAME, 'top-nav-info').click()
        time.sleep(1)
        wd.find_element(By.CLASS_NAME, 'account-tab-account').click()
        time.sleep(2)

        account = wd.find_element(By.ID, 'username')
        pwd = wd.find_element(By.ID, 'password')
        account.clear()
        pwd.clear()
        account.send_keys(douban_config.douban_id)
        pwd.send_keys(douban_config.douban_pwd)

        login = wd.find_element(By.XPATH, '//div[@class="account-form-field-submit "]//a[@class="btn btn-account btn-active"]')
        login.click() # 这里需要自己点一下验证码图片
        time.sleep(10)
    except NoSuchElementException:
        # 找不到该元素表示已登录
        print('已登录')
        time.sleep(1)
        wd.find_element(By.XPATH, '//a[@class="bn-more"]').click()

    # 打开个人主页
    wd.find_element(By.XPATH, '//a[@class="bn-more"]').click()
    wd.find_element(By.XPATH, '//div[@class="more-items"]//a').click()
    time.sleep(2)

    # 读过（我这里只有想读、读过的记录，没有在读）读过是第2项
    readed = wd.find_element(By.XPATH, '//div[@id="book"]//h2//a[@target="_blank"][2]')
    readed_site = readed.get_attribute('href')
    # print(readed_site)
    readed.click()

    # switch to new window 切换到新的窗口
    for handle in wd.window_handles:
        wd.switch_to.window(handle)
        if(wd.current_url == readed_site):
            break

    time.sleep(2)
    return wd

# 获取一页的数据内容
def get_one_page(wd, page_site):
    try:
        wd.get(page_site)
        time.sleep(2)
    except selenium.common.exceptions.TimeoutException:
        time.sleep(5)
        wd.get(page_site)

    pic_list = [] # 图书图片List
    pic_values = wd.find_elements(By.XPATH, '//li[@class="subject-item"]//img') # 图书图片
    for value in pic_values:
        pic_site = value.get_attribute('src')
        pic_list.append(pic_site)
    # print(pic_list)

    book_site_list = [] # 图片链接List
    book_name_list = [] # 图书名称List
    book_values = wd.find_elements(By.XPATH, '//li[@class="subject-item"]//div[@class="info"]//h2//a')
    for value in book_values:
        book_site = value.get_attribute('href')
        book_site_list.append(book_site)
        book_name = value.text
        book_name_list.append(book_name)
    # print(book_site_list)
    # print(book_name_list)

    author_list = [] # 作者信息List
    author_info = wd.find_elements(By.XPATH, '//li[@class="subject-item"]//div[@class="info"]//div[@class="pub"]')
    for value in author_info:
        author_list.append(value.text)
    # print(author_list)

    tags_list = [] # 标签List
    tag_values = wd.find_elements(By.XPATH, '//li[@class="subject-item"]//div[@class="info"]//div[@class="short-note"]//span[@class="tags"]')
    for value in tag_values:
        tags_list.append(value.text.split(':')[1].strip())
    # print(tags_list)

    date_list = [] # 读过日期List
    date_values = wd.find_elements(By.XPATH, '//li[@class="subject-item"]//div[@class="info"]//div[@class="short-note"]//span[@class="date"]')
    for value in date_values:
        date_list.append(value.text.split()[0])
    # print(date_list)

    rating_list = [] # 评分List（星星）
    rating_num_list = [] # 评分List（数字）
    rating_values = wd.find_elements(By.XPATH, '//li[@class="subject-item"]//div[@class="info"]//div[@class="short-note"]//span[1]')
    for value in rating_values:
        rate = value.get_attribute('class')[6] # str 5/4/3/2/1
        rating_num_list.append(int(rate))
        rating_list.append(douban_config.rating_dict[rate])
    # print(rating_list)
    # print(rating_num_list)

    comment_list = [] # 短评List
    comment_values =  wd.find_elements(By.XPATH, '//li[@class="subject-item"]//div[@class="info"]//div[@class="short-note"]//p[@class="comment"]')
    for value in comment_values:
        comment_list.append(value.text)
    # print(comment_list)

    num = len(book_name_list)
    data_list = []
    for i in range(num):
        data_list.append([book_name_list[i], book_site_list[i], author_list[i], tags_list[i], date_list[i], comment_list[i], rating_num_list[i], rating_list[i], pic_list[i]])
    return data_list

# 把记录数据List 全部存入数据库
def insert_into_db(db_file, table_name, data_list):
    columns = ['name', 'site', 'author', 'tags', 'date', 'comments', 'rating_num', 'rating', 'pic']

    sql = z_db.get_insert_sql_by_colum_names(table_name, columns)
    # print(sql) # insert into books (name, site, author, tags, date, comments, rating, pic) values (?, ?, ?, ?, ?, ?, ?, ?)
    z_db.insert_into_db(db_file, table_name, sql, data_list)

# 把记录数据List 全部存入Excel表
def write_to_excel(ex_file, sheet_name, data_list):
    if(os.path.exists(ex_file)):
        book = openpyxl.load_workbook(ex_file)
        if(sheet_name in book.sheetnames):
            sheet = book[sheet_name]
        else:
            sheet = book.create_sheet(sheet_name)
    else:
        book = openpyxl.Workbook()
        sheet = book.active
        sheet.title = sheet_name

    for data_line in data_list:
        sheet.append(data_line)
    book.save(ex_file)

# main
def main():
    site = douban_config.books_site
    wd = get_readed_wd(site)

    num_site_list = get_max_page_num(wd) # [最大页数, 模板url]

    max_page_num = num_site_list[0] # 最大页数
    # print(max_page_num) # 21
    model_site = num_site_list[1] # 模板url
    # print(model_site) # https://book.douban.com/people/[id]/collect?start=300&sort=time&rating=all&filter=all&mode=grid
    
    page_site_list = get_page_site_list(max_page_num, model_site) # 获取所有页数的链接List

    all_page_data_list = [] # 所有页面的数据信息
    for page_site in page_site_list:
        # print(page_site)
        print('------------第', page_site_list.index(page_site)+1, '页')
        one_page_data_list = get_one_page(wd, page_site)
        time.sleep(2)
        all_page_data_list = all_page_data_list + one_page_data_list
    # print(all_page_data_list)
    wd.quit()

    all_page_data_list.reverse() # 倒序

    # 读过的书籍数据信息写入数据库
    db_file = douban_config.db_file
    table_name = douban_config.books_table_name
    insert_into_db(db_file, table_name, all_page_data_list)
    time.sleep(3)

    # 读过的书籍数据信息写入Excel表格
    ex_file = douban_config.ex_file
    sheet_name = douban_config.book_sheet_name
    write_to_excel(ex_file, sheet_name, all_page_data_list)

    print('--------------------- Douban 已读书目 记录保存成功！')

# 清空数据库表数据
# z_db.delete_table(douban_config.db_file, douban_config.books_table_name)

# main()